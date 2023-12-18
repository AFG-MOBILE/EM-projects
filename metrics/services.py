from datetime import datetime, timedelta
import openpyxl
import pandas as pd
import commons_yape
import datadog
import nave
import linearb
import github_yape
from openpyxl.utils.dataframe import dataframe_to_rows
import functools
import cache_yape
import graphics
import newsletter
import subprocess
import powerpoint
import math


@cache_yape.daily_cache_clear
@functools.lru_cache(maxsize=None)
def getInfoReleases(month, year):
    # Data de início para listar releases
    inicio, fim = commons_yape.get_start_end_dates(year,month)
    periodo = f'inicio: {inicio} - fim: {fim}'
    
    # Data de início para listar releases
    data_inicial = datetime.strptime(f'{inicio} 00:00:00', '%d/%m/%y %H:%M:%S') #datetime.now()
    data_final = datetime.strptime(f'{fim} 23:59:59', '%d/%m/%y %H:%M:%S') #datetime.now()
    data_inicial_query_datadog = round(data_inicial.timestamp()) * 1000
    data_final_query_datadog = round(data_final.timestamp()) * 1000
    # Data de início para listar releases
    data_inicial = datetime.strptime(f'{inicio} 00:00:00', '%d/%m/%y %H:%M:%S') #datetime.now()
    data_final = datetime.strptime(f'{fim} 23:59:59', '%d/%m/%y %H:%M:%S') #datetime.now()
    data_inicial_query_datadog = round(data_inicial.timestamp()) * 1000
    data_final_query_datadog = round(data_final.timestamp()) * 1000
    
    releaseQA = datadog.getReleasesQA(data_inicial_query_datadog,data_final_query_datadog)
    releaseStaging = datadog.getReleasesStaging(data_inicial_query_datadog,data_final_query_datadog)
    releaseProduction = datadog.getReleasesProduction(data_inicial_query_datadog,data_final_query_datadog)
    release = releaseQA + releaseStaging + releaseProduction

    df_release = pd.DataFrame(release)
    # Filtrando os deploys com status "success"
    df_release_successful = df_release[df_release['status'] == 'success']
    # Agrupando por owner e contando os deploys de cada stage
    grouped1 = df_release_successful.groupby('owner')['stage'].value_counts().unstack(fill_value=0)
    # Adicionando uma nova coluna com a soma total de deploys por owner
    grouped1['Total Deploys'] = grouped1.sum(axis=1)
    grouped1 = grouped1[['Total Deploys', 'qa', 'staging', 'release']]
    return periodo, grouped1

@cache_yape.daily_cache_clear
@functools.lru_cache(maxsize=None)
def getInfoBugs(month, year):
    periodo, count_bugs_created, count_bugs_done, bugs, total_bugs = nave.getCardsByLabel(month,year, tuple(['bug']))
    # Agrupando por owner e contando os deploys de cada stage
    grouped1 = bugs.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped1['Bugs Done'] = grouped1.sum(axis=1)
    grouped1 = grouped1[['Bugs Done']]
    grouped2 = total_bugs.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped2['Bugs Total'] = grouped2.sum(axis=1)
    grouped2 = grouped2[['Bugs Total']]
    group = pd.merge(grouped2, grouped1, left_index=True, right_index=True, how='outer')
    return periodo, group

@cache_yape.daily_cache_clear
@functools.lru_cache(maxsize=None)
def getInfoDT(month, year):
    periodo, count_dt_created, count_dt_done, dts, total_dts = nave.getCardsByLabel(month,year, tuple(['debt', 'deuda']))
    # Agrupando por owner e contando os deploys de cada stage
    grouped1 = dts.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped1['DT Done'] = grouped1.sum(axis=1)
    grouped1 = grouped1[['DT Done']]

    grouped2 = total_dts.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped2['DT Total'] = grouped2.sum(axis=1)
    grouped2 = grouped2[['DT Total']]

    group = pd.merge(grouped2, grouped1, left_index=True, right_index=True, how='outer')
    return periodo, group

def getAllMetrics(month, year):
    print("obtendo metricas de cycle time")
    linearbMetrics = getCycleTime(month,year)
    print("obtendo metricas de releases")
    periodoR, releases = getInfoReleases(month,year)
    merged_grouped_diff = pd.merge(linearbMetrics, releases, left_index=True, right_index=True, how='outer') 
    print("obtendo metricas de bugs")
    periodoB, bugs = getInfoBugs(month,year)
    print("obtendo metricas de dts")
    periodoDT, deudas = getInfoDT(month,year)
    merged_grouped_diff = pd.merge(merged_grouped_diff, bugs, left_index=True, right_index=True, how='outer')
    merged_grouped_diff = pd.merge(merged_grouped_diff, deudas, left_index=True, right_index=True, how='outer')
    merged_grouped_diff.fillna(0, inplace=True)
    merged_grouped_diff['DT Done'] = merged_grouped_diff['DT Done'].apply(int)
    merged_grouped_diff['DT Total'] = merged_grouped_diff['DT Total'].apply(int)
    merged_grouped_diff['Bugs Done'] = merged_grouped_diff['Bugs Done'].apply(int)
    merged_grouped_diff['Bugs Total'] = merged_grouped_diff['Bugs Total'].apply(int)

    merged_grouped_diff['qa'] = merged_grouped_diff['qa'].apply(int)
    merged_grouped_diff['staging'] = merged_grouped_diff['staging'].apply(int)
    merged_grouped_diff['release'] = merged_grouped_diff['release'].apply(int)
    merged_grouped_diff['Total Deploys'] = merged_grouped_diff['Total Deploys'].apply(int)
    merged_grouped_diff['Total Deploys'] = merged_grouped_diff.apply(lambda x: f"{x['Total Deploys']} ({x['qa']} - {x['staging']} - {x['release']})", axis=1)
    merged_grouped_diff['total_release_deploy_raw'] = merged_grouped_diff['release'].apply(int)
    inicio, fim = commons_yape.get_start_end_dates(year,month)
    periodo = f"{inicio.replace('/','-')}_{fim.replace('/','-')}"
    merged_grouped_diff = merged_grouped_diff.reset_index()
    merged_grouped_diff = merged_grouped_diff.rename(columns={'index': 'owner'})

    owner_para_filtrar = ["owner_checkout","owner_crm","owner_gas","owner_insurance","owner_krossboarder-remesas","owner_marketplace","owner_promos","owner_tap2phone","owner_tipodecambio", "retail"]
    df_filtrado = merged_grouped_diff[merged_grouped_diff['owner'].isin(owner_para_filtrar)]
    df_filtrado = df_filtrado.rename(columns={'index': 'owner'})
    print("add dados por tribos")
    df_filtrado = appendTribe("retail", ["owner_gas","owner_marketplace","owner_promos"],df_filtrado)
    df_filtrado = appendTribe("financial", ["owner_tipodecambio","owner_krossboarder-remesas","owner_insurance"],df_filtrado)
    df_filtrado = appendTribe("negocios", ["owner_checkout","owner_tap2phone"],df_filtrado)
    df_filtrado = appendTribe("product cx", ["owner_crm"],df_filtrado)
    df_filtrado.fillna(0, inplace=True)
    df_raw = df_filtrado.copy()
    df_filtrado = df_filtrado[['owner','cycle time','coding','pickup','review','deploy','refactor','rework','work breakdown(newcode-refactor-rework)','Total Deploys','Bugs Total','Bugs Done','DT Total','DT Done']]
    df_raw.rename(columns={'cycle time (minutes)': 'cycletime_raw'}, inplace=True)
    df_raw.rename(columns={'Bugs Total': 'bugs_total_raw'}, inplace=True)
    df_raw.rename(columns={'DT Total': 'dt_total_raw'}, inplace=True)
    df_raw = df_raw[['owner','cycletime_raw','coding_raw', 'pickup_raw', 'review_raw', 'deploy_raw','refactor_raw','rework_raw','bugs_total_raw','dt_total_raw','total_release_deploy_raw']]
    print(df_raw)
    return periodo, df_filtrado, df_raw

def appendTribe(owner_tribe, filter, df):
    new_df = pd.DataFrame({
        'owner': [owner_tribe]
    })
    df_filtrado = df[df['owner'].isin(filter)]
    # Somar os valores das colunas 'Valor1' e 'Valor2'
    cycle_time_minutes = int(df_filtrado['cycle time (minutes)'].sum()/len(filter))
    coding_time_minutes = int(df_filtrado['coding_raw'].sum()/len(filter))
    pickup_time_minutes = int(df_filtrado['pickup_raw'].sum()/len(filter))
    review_time_minutes = int(df_filtrado['review_raw'].sum()/len(filter))
    deploy_time_minutes = int(df_filtrado['deploy_raw'].sum()/len(filter))
    new_df['Bugs Total'] = int(df_filtrado['Bugs Total'].sum())
    new_df['Bugs Done'] = int(df_filtrado['Bugs Done'].sum())
    new_df['DT Total'] = int(df_filtrado['DT Total'].sum())
    new_df['DT Done'] = int(df_filtrado['DT Done'].sum())
    new_df['cycle time (minutes)'] = cycle_time_minutes
    new_df['coding_raw'] = coding_time_minutes
    new_df['pickup_raw'] = pickup_time_minutes
    new_df['review_raw'] = review_time_minutes
    new_df['deploy_raw'] = deploy_time_minutes
    new_df['cycle time'] = f"{commons_yape.getBenchmark('cycleTime',cycle_time_minutes)} {commons_yape.formatTime(cycle_time_minutes)}"
    new_df['coding'] = f"{commons_yape.getBenchmark('codingTime',coding_time_minutes)} {commons_yape.formatTime(coding_time_minutes)}"
    new_df['pickup'] = f"{commons_yape.getBenchmark('pickupTime',pickup_time_minutes)} {commons_yape.formatTime(pickup_time_minutes)}"
    new_df['review'] = f"{commons_yape.getBenchmark('reviewTime',review_time_minutes)} {commons_yape.formatTime(review_time_minutes)}"
    new_df['deploy'] = f"{commons_yape.getBenchmark('deployTime',deploy_time_minutes)} {commons_yape.formatTime(deploy_time_minutes)}"
    new_df['qa'] = df_filtrado['qa'].sum()
    new_df['staging'] = df_filtrado['staging'].sum()
    new_df['release'] = df_filtrado['release'].sum()
    total = df_filtrado['release'].sum() + df_filtrado['staging'].sum() + df_filtrado['qa'].sum()
    new_df['Total Deploys'] = f"{total} ({df_filtrado['qa'].sum()} - {df_filtrado['staging'].sum()} - {df_filtrado['release'].sum()})"
    new_code = round(df_filtrado['new_code_raw'].sum()/len(filter),2)
    refactor = round(df_filtrado['refactor_raw'].sum()/len(filter),2)
    rework = round(100 - (new_code + refactor),2)
    new_df['refactor_raw'] = refactor
    new_df['rework_raw'] = rework
    new_df['refactor'] = f"{commons_yape.getBenchmark('refactor',refactor)} {refactor}%"
    new_df['rework'] = f"{commons_yape.getBenchmark('rework',rework)} {rework}%"
    new_df['work breakdown(newcode-refactor-rework)'] = f"{new_code}% - {refactor}% - {rework}%"
    df = pd.concat([df, new_df], ignore_index=True)
    return df

def comparar_colunas(df1, df2, colunas, sufixo_unicode='_unicode', sufixo_diferenca='_diff'):
    # Verificar se as colunas existem em ambos os DataFrames
    colunas_comuns = [col for col in colunas if col in df1.columns and col in df2.columns]
    df1.fillna(0, inplace=True)
    df2.fillna(0, inplace=True)

    df_diff = df1.copy()
    # Criar as novas colunas com análise no DataFrame 1
    for coluna in colunas_comuns:
        nome_coluna_unicode = coluna + sufixo_unicode
        nome_coluna_diferenca = coluna + sufixo_diferenca 
        print(coluna)
        if coluna == 'total_release_deploy_raw':
            df_diff[nome_coluna_unicode] = df_diff[coluna].compare(df2[coluna]).apply(
            lambda x: "⬆ +" if x['self'] < x['other'] else ("⬇ -" if x['self'] > x['other'] else "="), axis=1)
        else:
            df_diff[nome_coluna_unicode] = df_diff[coluna].compare(df2[coluna]).apply(
                lambda x: "⬆ -" if x['self'] > x['other'] else ("⬇ +" if x['self'] < x['other'] else "="), axis=1)
        df_diff[nome_coluna_diferenca] = df_diff[coluna].compare(df2[coluna]).apply(
            lambda x: abs(x['self'] - x['other']), axis=1)
    df_diff = df_diff[['owner', 'cycletime_raw_unicode',
       'cycletime_raw_diff', 'coding_raw_unicode', 'coding_raw_diff',
       'pickup_raw_unicode', 'pickup_raw_diff', 'review_raw_unicode',
       'review_raw_diff', 'deploy_raw_unicode', 'deploy_raw_diff',
       'refactor_raw_unicode', 'refactor_raw_diff', 'rework_raw_unicode',
       'rework_raw_diff', 'refactor_raw', 'rework_raw','bugs_total_raw_unicode','dt_total_raw_unicode','bugs_total_raw_diff','dt_total_raw_diff','total_release_deploy_raw_unicode','total_release_deploy_raw_diff']]
    
    df_diff['refactor_raw_diff'] = df_diff['refactor_raw_diff'].round(2)
    df_diff['refactor_raw_diff'] = df_diff['refactor_raw_diff'].astype(str)
    df_diff['rework_raw_diff'] = df_diff['rework_raw_diff'].round(2)
    df_diff['rework_raw_diff'] = df_diff['rework_raw_diff'].astype(str)
    df_diff['cycletime_diff'] = df_diff['cycletime_raw_unicode'] + " " + df_diff['cycletime_raw_diff'].apply(commons_yape.formatTime)
    df_diff['coding_diff'] = df_diff['coding_raw_unicode'] + " " +  df_diff['coding_raw_diff'].apply(commons_yape.formatTime)
    df_diff['pickup_diff'] = df_diff['pickup_raw_unicode'] + " " +  df_diff['pickup_raw_diff'].apply(commons_yape.formatTime)
    df_diff['review_diff'] = df_diff['review_raw_unicode'] + " " +  df_diff['review_raw_diff'].apply(commons_yape.formatTime)
    df_diff['deploy_diff'] = df_diff['deploy_raw_unicode'] + " " +  df_diff['deploy_raw_diff'].apply(commons_yape.formatTime)
    df_diff['refactor_diff'] = df_diff['refactor_raw_unicode'] + " " + df_diff['refactor_raw_diff'] + "%"
    df_diff['rework_diff'] = df_diff['rework_raw_unicode'] + " " +  df_diff['rework_raw_diff'] + "%"
    df_diff['deploy_diff'] = df_diff['deploy_diff'].fillna("=")
    df_diff['bugs_total_diff'] = df_diff['bugs_total_raw_unicode'] + " " + df_diff['bugs_total_raw_diff'].round(2).astype(str)
    df_diff['bugs_total_diff'] = df_diff['bugs_total_diff'].fillna("=")
    df_diff['dt_total_diff'] = df_diff['dt_total_raw_unicode'] + " " + df_diff['dt_total_raw_diff'].round(2).astype(str)
    df_diff['dt_total_diff'] = df_diff['dt_total_diff'].fillna("=")
    df_diff['total_release_deploy_diff'] = df_diff['total_release_deploy_raw_unicode'] + " " + df_diff['total_release_deploy_raw_diff'].round(2).astype(str)
    df_diff['total_release_deploy_diff'] = df_diff['total_release_deploy_diff'].fillna("=")
    df_diff = df_diff[['owner', 
       'cycletime_diff', 'coding_diff',
       'pickup_diff', 'review_diff', 'deploy_diff', 
       'refactor_diff', 'rework_diff','bugs_total_diff','dt_total_diff','total_release_deploy_diff']]
    return df_diff

def postReleasesInLinearB(days=3):

    data_inicial = datetime.now() - timedelta(days=days)
    data_final = datetime.now()
    data_inicial_query_datadog = round(data_inicial.timestamp()) * 1000
    data_final_query_datadog = round(data_final.timestamp()) * 1000
    
    releaseQA = datadog.getReleasesQA(data_inicial_query_datadog,data_final_query_datadog)
    releaseStaging = datadog.getReleasesStaging(data_inicial_query_datadog,data_final_query_datadog)
    releaseProduction = datadog.getReleasesProduction(data_inicial_query_datadog,data_final_query_datadog)
    releases = releaseQA + releaseStaging + releaseProduction
    df_release = pd.DataFrame(releases)
    df_release_successful = df_release[df_release['status'] == 'success']
    executes = 0
    for index, row in df_release_successful.iterrows():
        status = linearb.execute_linearb_post(row['repo_name'],row['stage'], row['commit_sha'], row['timestamp'], row['owner'])
        if status:
            executes += 1
    return f'Foram executadas com sucesso {executes}/{len(df_release_successful)} para o periodo de {days} dias atrás'

def checkNewServicesWithoutOwners():
    # obter todos os repositorios do github
    all_repositories = github_yape.get_all_repositories_github()
    #  verifica no github os repositorios que nao possuem um owner 
    repositories_without_owners = github_yape.check_repositories_without_owners_in_github(all_repositories)
    if len(repositories_without_owners) > 0:
        df_repositories_without_owners = pd.DataFrame(repositories_without_owners)
        df_repositories_without_owners.columns = ['Repositorios']
        print("***********************************************************************")
        print(f'Repositorios sem owner no github: \n {df_repositories_without_owners}')
        print("***********************************************************************")
    else:
        print("***********************************************************************")
        print(f'Não foram encontrados repositorios sem owner no github')
        print("***********************************************************************")

    # Verifica se falta add um repositorio que possua um owner mas ainda nao está em algum serviço do linearb
    repositories_without_services = linearb.check_repositories_in_linearb(all_repositories)
    if len(repositories_without_services) > 0:
        df_repositories_without_services = pd.DataFrame(repositories_without_services)
        df_repositories_without_services.columns = ['Repositorios']
        print("***********************************************************************")
        print(f'Repositorios no github que ainda nao foram alocados no linearB: \n {df_repositories_without_services}')
        print("***********************************************************************")
    else:
        print("***********************************************************************")
        print(f'Não foram encontrados repositorios no github que ainda nao foram alocados no linearB')
        print("***********************************************************************")

def getCycleTime(month, year):
    services = linearb.get_services()
    metrics = []
    for owner, serviceId in services.items():
        owner_metrics = linearb.get_measurements(tuple([serviceId]), owner, month, year)
        metrics.append(owner_metrics)

    # teams = linearb.get_teams()
    # for owner, teamId in teams.items():
    #     owner_metrics = linearb.get_measurementsByTeam(tuple([teamId]), owner, month, year)
    #     metrics.append(owner_metrics)

    df_metrics =  pd.DataFrame(metrics)
    df_metrics.fillna(0, inplace=True)
    df_metrics = df_metrics.set_index('owner')
    return df_metrics
    
def getSprintInfos(data_inicio, data_fim, owner, label):
    data_dashboard = nave.getDashBoardData(owner,data_inicio,data_fim)
    nave.getSpotGraph(data_inicio, data_fim, label,data_dashboard)

def getBugsGraph(data_inicio, data_fim, owner):
    data_dashboard = nave.getDashBoardData(owner,data_inicio,data_fim)
    data_dashboard_filter = nave.__filter_tasks_by_label(data_dashboard.copy(), 'Bug')
    graphics.plot_bugs_restantes_por_mes(data_dashboard_filter)
    # nave.getBugsCreatedVsFinished(data_inicio, data_fim,data_dashboard)

def getDeudasTecnicas(data_inicio, data_fim, owner):
    data_dashboard = nave.getDashBoardData(owner,data_inicio,data_fim)
    nave.getDTCreatedVsFinished(data_inicio, data_fim,data_dashboard)

def getCycleTimeTribu(month, year):
    metrics = []
    tribes = linearb.get_services_by_tribe()
    for owner, services in tribes.items():
        owner_metrics = linearb.get_measurements(list(services), owner, month, year)
        metrics.append(owner_metrics)
    return metrics

def checkMetricsByMonth(months, year):
    YEAR = year

    monthsText = {1: 'Enero',2: 'Febrero',3: 'Marzo',4: 'Abril',5: 'Mayo',6: 'Junio',7: 'Julio',8: 'Agosto',9: 'Septiembre',10: 'Octubre',11: 'Noviembre',12: 'Diciembre'}
    last_index = len(months)-1
    second_last_index = len(months)-2
    # Create a new Excel workbook
    wb_new = openpyxl.Workbook()
    del wb_new["Sheet"]  # remove default sheet
    wb_new_raw = openpyxl.Workbook()
    del wb_new_raw["Sheet"]  # remove default sheet
    list_metrics_raw = {}
    list_metrics = {}
    for month in months:
        periodo, metrics, metrics_raw = getAllMetrics(month,YEAR)
        # Add data to a new sheet in the Excel workbook
        list_metrics_raw[month] = metrics_raw
        list_metrics[month] = metrics
    
    colunas_para_comparar = ['cycletime_raw', 'coding_raw', 'pickup_raw', 'review_raw', 'deploy_raw', 'refactor_raw', 'rework_raw', 'bugs_total_raw', 'dt_total_raw','total_release_deploy_raw']
    df = comparar_colunas(list_metrics_raw[months[second_last_index]],list_metrics_raw[months[last_index]], colunas_para_comparar)
    merged_grouped_diff = pd.merge(list_metrics[months[last_index]], df, left_index=True, right_index=True, how='outer')
    merged_grouped_diff = merged_grouped_diff.drop('owner_y', axis=1)
    merged_grouped_diff = merged_grouped_diff.rename(columns={'owner_x': 'owner'})

    
    list_metrics[months[last_index]] = merged_grouped_diff
    for month in months:
        ws_month = wb_new.create_sheet(title=monthsText[month])
        
        for r_idx, row in enumerate(dataframe_to_rows(list_metrics[month], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_month.cell(row=r_idx, column=c_idx, value=value)
        
        ws_month_raw = wb_new_raw.create_sheet(title=monthsText[month])
        for r_idx, row in enumerate(dataframe_to_rows(list_metrics_raw[month], index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_month_raw.cell(row=r_idx, column=c_idx, value=value)

    # Save to an Excel file
    filename_monthly = "/Users/alexfrisoneyape/Development/EM-projects/metrics/metricas.xlsx"
    filename_monthly_analysis = "/Users/alexfrisoneyape/Development/EM-projects/metrics/metricas_analysis.xlsx"
    wb_new.save(filename_monthly)
    wb_new_raw.save(filename_monthly_analysis)
    subprocess.run(['open', '-a', 'Microsoft Excel', '/Users/alexfrisoneyape/Development/EM-projects/metrics/metricas.xlsx' ])
    return list_metrics[months[last_index]]

def generateNewsletter(template,news):
    html = newsletter.preencher_template_informe_semanal(template,news)
    with open('/Users/alexfrisoneyape/Development/EM-projects/metrics/informe_mensal.html', 'w') as file:
        file.write(html)
        
    return html

def createSlideForShowcase(path, nome_da_aba):

    # Carregar a aba específica em um DataFrame
    df = pd.read_excel(path, sheet_name=nome_da_aba)
    df.fillna(0, inplace=True)
    df['cycletime'] = df['cycle time'] + "\n (" + df['cycletime_diff'] + ")"
    df = df.fillna("=")
    df['dt'] = df['DT Total'].astype(str) + " / " + df['DT Done'].astype(str) + "\n (" + df['dt_total_diff'].astype(str) + ")"
    df['bugs'] = df['Bugs Total'].astype(str) + " / " + df['Bugs Done'].astype(str) + "\n (" + df['bugs_total_diff'].astype(str) + ")"
    df['deploys'] = df['Total Deploys'].astype(str) + "\n (" + df['total_release_deploy_diff'].astype(str) + ")"
    df['wb'] = df['work breakdown(newcode-refactor-rework)'] + "\n (" + df['refactor_diff'].astype(str) + " / " + df['rework_diff'].astype(str) + ")"
    
    #Removendo as coluna de não serão utilizadas
    df = df[['owner','cycletime','dt','bugs','deploys','wb']]
    
    #Criando o dicionario com as colunas
    data = {}
    # Iterando sobre o DataFrame usando iterrows()
    for indice, linha in df.iterrows():
        valor_coluna_A = linha['owner']  # Obtendo o valor da coluna 'A'
        
        # Iterando sobre as outras colunas (exceto 'A') usando items()
        for coluna, valor in linha.items():
            if coluna != 'owner':  # Ignorando a coluna 'A'
                chave = f"{valor_coluna_A}_{coluna}"  # Criando a chave concatenada
                data[chave] = valor  # Armazenando no dicionário
    # print(list(data.keys()))
    path_new_presentation = f'/Users/alexfrisoneyape/Development/EM-projects/{nome_da_aba} - Deck de Metricas.pptx'
    powerpoint.replace_hashtags_slide('/Users/alexfrisoneyape/Development/EM-projects/assets/Template - Deck de Metricas.pptx',data, path_new_presentation)
    subprocess.run(['open', path_new_presentation])
    #Enviar o caminho da apresentacao
    return 0


@cache_yape.daily_cache_clear
@functools.lru_cache(maxsize=None)
def getInfoDevOps(month, year):
    periodo, count_accesos_created, count_accesos_done, accesos, total_acesos = nave.getCardsByLabel(month,year, tuple(['accesos']))
    # Agrupando por owner e contando os acessos de cada stage
    grouped1 = accesos.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped1['Accesos Done'] = grouped1.sum(axis=1)
    grouped1 = grouped1[['Accesos Done']]

    periodo, count_apoyos_created, count_apoyos_done, apoyos, total_apoyos = nave.getCardsByLabel(month,year, tuple(['apoyo']))
    # Agrupando por owner e contando os apoios de cada stage
    grouped2 = apoyos.groupby('Owner')['Task name'].value_counts().unstack(fill_value=0)
    grouped2['Apoyos Done'] = grouped2.sum(axis=1)
    grouped2 = grouped2[['Apoyos Done']]

    group = pd.merge(grouped2, grouped1, left_index=True, right_index=True, how='outer')
    return periodo, group
